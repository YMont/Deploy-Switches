

''' 20/09/24-
# 1.
# modify code
# line 56 & 65
# line 93 & 94
# open(r"C:/Users/Edgar Ho/Desktop/%s_Pool.txt"%file_name,"a+")
# to 
# user_path = os.path.join(os.environ['USERPROFILE'],'Desktop')
# >>make program can use in any windows platform 
'''
# 20/10/22-

import re , os
from openpyxl import load_workbook
from logging import NOTSET , DEBUG , INFO , WARNING , ERROR , CRITICAL

counter_loop_row = 0

def user_path():
    var_path = os.path.join(os.environ['USERPROFILE'],'Desktop')
    return var_path # C:\Users\Edgar Ho\Desktop

def get_excel_infor(var_excel_path):
    wb = load_workbook(r"%s\DHCP_Pool.xlsx"%var_excel_path)
    ws = wb["sheet"]
    max_r = ws.max_row
    max_c = ws.max_column
    # print(wb , ws , max_r , max_c)
    return wb , ws , max_r , max_c
    # (<openpyxl.workbook.workbook.Workbook object at 0x09523160>, 
    # <Worksheet "sheet">,
    # 11,
    # 8)

def write_txt(write_final):
    a, b, c, d = get_excel_infor(user_path())
    # wb =a
    # ws = b
    max_r = c
    # max_c = d

    path = user_path()

    with open(r"%s\DHCP_Pool.txt"%(path),"a+") as ff : # write & append result in txt file    
        if counter_loop_row == max_r+1: #couter = 11 , max_row = 1
            # print(counter_loop_row)
            ff.write(write_final.strip("\n")) # last 
        else:
            ff.write(write_final+"\n")

def create_txt():
    path = user_path()
    with open(r"%s\DHCP_Pool.txt"%(path),"a+") as var_Initext : # initial file cursor,starting in beginning
        var_Initext.truncate(0) # initial action 

def ascii(input_row,input_column):
    a, b, c, d = get_excel_infor(user_path())
    # wb = a
    ws = b
    # max_r = c
    # max_c = d
    # print(max_c,max_r)
    
    res_SN = ws.cell(row=input_row, column=input_column).value
    char1 = res_SN
    after = "00"
    i=0

    for x in char1:
        tohex = hex(ord(x))
        res = tohex[2:4]
        after += res
        if i % 2 == 0  and i < len(char1)-1:
            after += "."
        i += 1
        if i == len(char1):
            # print("char:",char1)
            # print("char to hex:",after)
            return after

def dhcp_pool(): # staic use excel format , just follow it.
    a, b, c, d = get_excel_infor(user_path())
    # wb = a
    ws = b
    max_r = c
    max_c = d
    print(max_r,max_c)
    global counter_loop_row
    # for i in range(2,max_r+1): # excel-row    start 2,end (max_r)+1
    for i in range(2,11): # excel-row    start 2,end (max_r)+1
        counter_loop_row = i
        # print(counter_loop_row)
        if i is None:
            continue
        else:
            for x in range(3,max_c+1): # excel-column start 3,end (max_c)+1
                if x is None : # if meet "None" in excel space,loop skip,next                                                             
                    continue
                elif x == 3 :
                    res_1 = ws.cell(row=i,column=x).value
                    res_final = "ip dhcp pool"+" "+str(res_1) 
                    # print(res_final) # ip dhcp pool JAE2419176C
                    write_txt(res_final)

                elif x == 4 :
                    res_1 = ws.cell(row=i,column=x).value
                    res_2 = ws.cell(row=i,column=x+1).value
                    res_final = " "+"host"+" "+str(res_1)+" "+str(res_2) 
                    # print(res_final) # host xxx.xxx.xxx.xxx 255.255.255.0
                    write_txt(res_final)

                elif x == 5 :
                    # print("loop here.")
                    continue

                elif x == 6 :
                    res_1 = ascii(i,x) # call function transfer S/N to ASCii
                    res_final = " "+"client-identifier"+" "+str(res_1)
                    # print(res_final) # client-identifier 0046.4f43.3233.4647.5831.6649
                    write_txt(res_final) 

                elif x == 7 :
                    res_1 = ws.cell(row=i,column=x).value
                    res_final = " "+"option 150 ip"+" "+str(res_1) 
                    # print(res_final) # option 150 ip 192.168.xxx.xxx
                    write_txt(res_final)

                elif x == 8 :
                    res_1 = ws.cell(row=i,column=x).value
                    res_final = " "+"option 67 ascii"+" "+ "/"+str(res_1).strip("\n")
                    # print(res_final) # option 67 ascii /r1-confg
                    write_txt(res_final)
    return counter_loop_row
    
def main1():
    create_txt()
    dhcp_pool()

if __name__ == "__main__":
    try:
        main1()
    except ValueError as vale:
        print("Value:",vale)
    except StopIteration as sto:
        print("Stop:",sto)
    except Exception as exce:
        print("Error:",exce)
